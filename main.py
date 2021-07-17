import os
from openpyxl import Workbook, load_workbook


def extract_data(the_workbook, file_name):

    #with open(file_name, 'w+') as g:

    for sheet in the_workbook.sheetnames:
        current_ws = the_workbook[sheet]
        # Get unfiltered data in list l
        l = list()

        week_start = current_ws['A3'].value[current_ws['A3'].value.upper().find('FROM: ')+6:current_ws['A3'].value.upper().find('FROM: ')+6+10:1]
        week_end = current_ws['A3'].value[current_ws['A3'].value.upper().find('TO: ')+4::1]

        for row in current_ws.iter_rows(min_row=7, max_col=3, max_row=len(list(current_ws.rows)), values_only=True):
            l.append(row)

        for i in l:
            if ((str(i[0]).upper().find('TOTAL')) == -1 and (str(i[1]).upper().find('TOTAL')) == -1):
                l2.append((i[0], i[2], sheet, file_name, week_start, week_end))
    print(f'\nExtracted data from file {file_name}: \n{l2}')



def print_hi(name):
    print(f'{name}')


if __name__ == '__main__':
    print_hi('\nHello, my name is Svetlana Galuzinschii. This application will output the \'data_extract.txt\' file in the \'output-data\' directory in the application root. '
             '\nThe output you requested in the assignment is also displayed in the console output as two lists of tuples, '
             '\neach tuple corresponds to records discovered in a given Excel file. '
             '\nThe provided input data (Excel files) is stored in \'input-data\' folder in the application root.')

    basedir = os.path.abspath(os.path.dirname(__file__))
    data_files = os.listdir(basedir+'/input-data')
    print(f'\nProcessed data files are {data_files}')
    with open('output-data/data_extract.txt', 'w') as d:
        l2 = list()  # l2 is the filtered list with no 'total' word in the carload name
        for f in data_files:
            extract_data(load_workbook(basedir+'/input-data/'+f),f)
        d.write(str(l2))



